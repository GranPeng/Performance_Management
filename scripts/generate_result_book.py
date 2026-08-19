#!/usr/bin/env python3
"""生成绩效结果说明书 V03（只读 Base，写本地 XLSX）。

用途：将 Canonical Base 中的 Target、Actual、Performance_Result 及主数据关联还原为
可供 HR 人工复核的“录入侧 → 输出侧”说明书。按部门分 Sheet，每名员工一行并展示
负责渠道；V04 评分标准文本来自本地权威
结构化提取 JSON；不修改任何 Base 记录。失败会写出可追溯的错误日志。
"""
from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_BASE_TOKEN = "FCxObLU6yao5jgsciZfcWHKwnjh"
TABLES = {
    "Performance_Result": "tbl6tFtVKExFUTWo",
    "Actual": "tbli9VhcUFjVDeNd",
    "Target": "tblydZkf17kmzrO0",
    "Employee": "tblc59aB4EnSxkQv",
    "Position": "tbldzvsg9Op6pK29",
    "Metric": "tbldKtdIVv8nnTyX",
    "Organization": "tblc6rU0d2bHMVnZ",
    "Channel": "tblqOGJknsD2H3bt",
}
V04_RULES = ROOT / "data/output/绩效框架结构化规则.json"
OUT = ROOT / "data/output/绩效结果说明书_20260818_V03.xlsx"
ERROR_LOG = ROOT / "data/output/绩效结果说明书_20260818_V03_生成错误日志.json"
EXCEPTION_LOG = ROOT / "data/output/绩效结果说明书_20260818_V03_异常日志.json"
CLI = shutil.which("lark-cli") or str(Path.home() / ".local/bin/lark-cli")


def now_utc() -> str:
    return datetime.now(timezone.utc).isoformat()


def cli(base_token: str, args: list[str]) -> dict[str, Any]:
    """调用只读 lark-cli；非成功响应显式失败，绝不吞错。"""
    proc = subprocess.run(
        [CLI, "base", *args, "--base-token", base_token, "--as", "user"],
        text=True,
        capture_output=True,
        timeout=300,
    )
    try:
        response = json.loads(proc.stdout)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"lark-cli 返回非 JSON：{proc.stdout[-500:]} stderr={proc.stderr[-500:]}") from exc
    if proc.returncode != 0 or not response.get("ok"):
        raise RuntimeError(json.dumps(response, ensure_ascii=False))
    return response["data"]


def records(base_token: str, table: str, fields: list[str]) -> list[dict[str, Any]]:
    """按分页读取指定字段；保留 Base record_id 以确保关联可追踪。"""
    result: list[dict[str, Any]] = []
    offset = 0
    while True:
        args = [
            "+record-list", "--table-id", TABLES[table], "--limit", "200",
            "--format", "json", "--offset", str(offset),
        ]
        for field in fields:
            args += ["--field-id", field]
        data = cli(base_token, args)
        returned_fields = data.get("fields", [])
        ids = data.get("record_id_list", [])
        values = data.get("data", [])
        if len(ids) != len(values):
            raise RuntimeError(f"{table} 返回 record_id 与数据行数不一致：{len(ids)} != {len(values)}")
        for record_id, row_values in zip(ids, values):
            row = dict(zip(returned_fields, row_values))
            row["_record_id"] = record_id
            result.append(row)
        if not data.get("has_more"):
            return result
        if not ids:
            raise RuntimeError(f"{table} 声称存在下一页但本页无记录，已停止以防无限循环")
        offset += len(ids)


def link_id(value: Any) -> str | None:
    if isinstance(value, list) and value and isinstance(value[0], dict):
        return value[0].get("id")
    return None


def link_ids(value: Any) -> list[str]:
    """读取多值关联 record_id，保留 Base 返回顺序，不按名称猜测关系。"""
    if not isinstance(value, list):
        return []
    return [item["id"] for item in value if isinstance(item, dict) and item.get("id")]


def text(value: Any, default: str = "") -> str:
    if value is None:
        return default
    if isinstance(value, bool):
        return "是" if value else "否"
    return str(value)


def numeric(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, str) and value.endswith("%"):
        try:
            return float(value[:-1]) / 100
        except ValueError:
            return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def display_number(value: Any, decimals: int = 4) -> Any:
    number = numeric(value)
    if number is None:
        return ""
    return round(number, decimals)


def display_percent(value: Any) -> str:
    number = numeric(value)
    return "" if number is None else f"{number * 100:.2f}%"


def d013_tier_percentage(score_total: Any) -> float | None:
    """D-013 左闭右开档位；150 分及以上按 D-009 使用最高档。"""
    score = numeric(score_total)
    if score is None:
        return None
    if score < 60:
        return 0.8
    if score < 80:
        return 0.9
    if score < 100:
        return 1.0
    return 1.1


def v04_index() -> tuple[dict[tuple[str, str], dict[str, Any]], dict[str, Any]]:
    payload = json.loads(V04_RULES.read_text(encoding="utf-8"))
    index: dict[tuple[str, str], dict[str, Any]] = {}
    for position in payload["positions"]:
        for metric in position["metrics"]:
            index[(position["position"], metric["name"])] = {
                "source_sheet": position["source_sheet"],
                "source_sheet_code": position["source_sheet_code"],
                "source_cell": metric["source_cell"],
                "standard": metric["scoring_standard"],
                "formula": metric.get("calculation_formula") or "",
                "scoring_type": metric["scoring_type"],
                "source_sha256": payload["source"]["sha256"],
            }
    return index, payload["source"]


def explanation(row: dict[str, Any]) -> str:
    """只解释本行已存储的计算结果，不重新写业务规则。"""
    standard = row["V04评分标准原文"]
    rate = numeric(row["达成率"])
    final_score = row["最终得分"]
    auto_score = row["自动得分"]
    manual_score = row["人工分"]
    scoring_type = row["评分类型"]
    if scoring_type in {"定性等级型", "奖惩制"} and manual_score in {"", "待人工分"}:
        return "该指标须负责人按左侧 V04 原文录入人工分；当前未填人工分，输出标记为待人工分，不以 0 替代。"
    if scoring_type in {"定性等级型", "奖惩制"}:
        return f"该指标按左侧 V04 原文人工判定；本行人工分={manual_score}，最终得分={final_score}。"
    if rate is None:
        return f"未形成可计算达成率（目标或实际值缺失/目标为零）；按左侧 V04 原文，当前自动得分={auto_score or '空'}，最终得分={final_score or '空'}。"
    if rate >= 1:
        return f"达成率 {display_percent(rate)} ≥100%；对应左侧 V04 原文的达成/封顶规则，自动得分={auto_score}，最终得分={final_score}。"
    if rate >= 0.8:
        return f"达成率 {display_percent(rate)} 落入 V04 原文 80%~100% 区间；自动得分={auto_score}，最终得分={final_score}（以左侧原文具体阈值为准）。"
    if rate >= 0.6:
        return f"达成率 {display_percent(rate)} 落入 V04 原文低于 80% 的阈值区间；自动得分={auto_score}，最终得分={final_score}（以左侧原文具体阈值为准）。"
    return f"达成率 {display_percent(rate)} 低于 V04 原文常见最低阈值；自动得分={auto_score}，最终得分={final_score}。"


def fetch_data(base_token: str) -> dict[str, list[dict[str, Any]]]:
    return {
        "results": records(base_token, "Performance_Result", [
            "Result_ID", "Period", "Employee_ID", "Metric_ID", "Target_ID", "Actual_ID",
            "Target_Value_Snapshot", "Achievement_Rate", "Auto_Score", "Manual_Score",
            "Final_Score", "Weight", "Weighted_Score", "Monthly_Total",
            "Commission_Tier_Level", "Commission_Base", "Commission_Ratio", "Commission_Amount",
            "Perf_Salary_Snapshot",
            "Is_Exempt", "Review_Status", "Status", "Source", "Note",
        ]),
        "actuals": records(base_token, "Actual", [
            "Actual_ID", "Actual_Value", "Employee_ID", "Metric_ID", "Period", "Source", "Status", "Note",
        ]),
        "targets": records(base_token, "Target", [
            "Target_ID", "Target_Value", "Metric_ID", "Period", "Source", "Status", "Note",
        ]),
        "employees": records(base_token, "Employee", [
            "Employee_ID", "Name", "Position_ID", "Org_ID", "Responsible_Channel_IDs", "Perf_Participate_Status", "Status",
        ]),
        "positions": records(base_token, "Position", ["Position_ID", "Position_Name", "Job_Family", "Org_ID", "Status"]),
        "metrics": records(base_token, "Metric", [
            "Metric_ID", "Metric_Name", "Scoring_Type", "Dimension", "Weight", "Position_ID",
            "Source_Cell", "Scoring_Standard_Text", "Status",
        ]),
        "organizations": records(base_token, "Organization", ["Org_ID", "Org_Name", "Status"]),
        "channels": records(base_token, "Channel", ["Channel_ID", "Channel_Name", "Status"]),
    }


def make_rows(data: dict[str, list[dict[str, Any]]], v04: dict[tuple[str, str], dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    actual_by_record = {r["_record_id"]: r for r in data["actuals"]}
    target_by_record = {r["_record_id"]: r for r in data["targets"]}
    employee_by_record = {r["_record_id"]: r for r in data["employees"]}
    metric_by_record = {r["_record_id"]: r for r in data["metrics"]}
    position_by_record = {r["_record_id"]: r for r in data["positions"]}
    org_by_record = {r["_record_id"]: r for r in data["organizations"]}
    channel_by_record = {r["_record_id"]: r for r in data["channels"]}
    # D-013 的中控人数以同期间、同负责渠道的中控结果去重统计；不从说明文本猜测。
    d013_people_by_period_channel: dict[tuple[str, str], set[str]] = defaultdict(set)
    for candidate in data["results"]:
        if candidate.get("Source") != "SIMULATED_T21_D013_D014":
            continue
        try:
            candidate_note = json.loads(candidate.get("Note") or "{}")
        except json.JSONDecodeError:
            continue
        if str(candidate_note.get("t21_case", "")).startswith("D013"):
            channel_id = text(candidate_note.get("channel_record_id"))
            employee_record_id = link_id(candidate.get("Employee_ID"))
            if channel_id and employee_record_id:
                d013_people_by_period_channel[(text(candidate.get("Period")), channel_id)].add(employee_record_id)

    input_rows: list[dict[str, Any]] = []
    output_rows: list[dict[str, Any]] = []
    for result in data["results"]:
        employee = employee_by_record.get(link_id(result.get("Employee_ID")), {})
        metric = metric_by_record.get(link_id(result.get("Metric_ID")), {})
        position = position_by_record.get(link_id(employee.get("Position_ID")), {})
        organization = org_by_record.get(link_id(employee.get("Org_ID")), {})
        actual = actual_by_record.get(link_id(result.get("Actual_ID")), {})
        target = target_by_record.get(link_id(result.get("Target_ID")), {})
        position_name = text(position.get("Position_Name"), "未关联岗位")
        metric_name = text(metric.get("Metric_Name"), "未关联指标")
        channel_names = [
            text(channel_by_record[channel_id].get("Channel_Name"))
            for channel_id in link_ids(employee.get("Responsible_Channel_IDs"))
            if channel_id in channel_by_record
        ]
        channel_display = "、".join(name for name in channel_names if name) or "—"
        rule = v04.get((position_name, metric_name))
        if rule is None:
            # Base 已保存的标准文本是 V04 提取落库字段；仍保留缺失标记，禁止无提示替换。
            rule = {
                "source_sheet": "未匹配",
                "source_sheet_code": "未匹配",
                "source_cell": text(metric.get("Source_Cell"), ""),
                "standard": text(metric.get("Scoring_Standard_Text"), "【V04 原文未匹配：需核查指标/岗位映射】"),
                "formula": "",
                "scoring_type": text(metric.get("Scoring_Type")),
            }
        target_value = target.get("Target_Value")
        if target_value in (None, ""):
            target_value = result.get("Target_Value_Snapshot")
        rate = result.get("Achievement_Rate")
        d013_review = {"绩效工资基数": "", "档位百分比": "", "绩效工资部分金额": "", "中控人数": "", "渠道GSV": "", "GSV提成部分金额": "", "复合收入（两部分合计）": ""}
        if result.get("Source") == "SIMULATED_T21_D013_D014":
            try:
                d013_note = json.loads(result.get("Note") or "{}")
            except json.JSONDecodeError:
                d013_note = {}
            if str(d013_note.get("t21_case", "")).startswith("D013"):
                channel_id = text(d013_note.get("channel_record_id"))
                middle_control_count = len(d013_people_by_period_channel[(text(result.get("Period")), channel_id)])
                salary = numeric(result.get("Perf_Salary_Snapshot"))
                tier_percentage = d013_tier_percentage(result.get("Monthly_Total"))
                gsv = numeric(result.get("Commission_Base"))
                gsv_commission = numeric(result.get("Commission_Amount"))
                salary_part = salary * tier_percentage if salary is not None and tier_percentage is not None else None
                composite_income = salary_part + gsv_commission if salary_part is not None and gsv_commission is not None else None
                d013_review = {
                    "绩效工资基数": display_number(salary),
                    "档位百分比": display_percent(tier_percentage),
                    "绩效工资部分金额": display_number(salary_part),
                    "中控人数": middle_control_count or "",
                    "渠道GSV": display_number(gsv),
                    "GSV提成部分金额": display_number(gsv_commission),
                    "复合收入（两部分合计）": display_number(composite_income),
                }
        common = {
            "期间": text(result.get("Period")), "姓名": text(employee.get("Name"), "未关联员工"),
            "部门": text(organization.get("Org_Name"), "未关联部门"), "岗位": position_name, "渠道": channel_display,
            "员工ID": text(employee.get("Employee_ID")), "指标ID": text(metric.get("Metric_ID")),
            "指标名称": metric_name, "评分类型": text(metric.get("Scoring_Type")),
            "Result记录ID": result["_record_id"], "结果ID": text(result.get("Result_ID")),
        }
        input_row = {
            **common,
            "Target记录ID": target.get("_record_id", ""), "目标ID": text(target.get("Target_ID")),
            "目标值": display_number(target_value), "Target来源批次": text(target.get("Source")),
            "Actual记录ID": actual.get("_record_id", ""), "实际ID": text(actual.get("Actual_ID")),
            "实际值": display_number(actual.get("Actual_Value")), "Actual来源批次": text(actual.get("Source")),
            "达成率": display_percent(rate), "数据完整性": "完整" if target and actual else "关联缺失/不适用：详见ID空值",
        }
        output_row = {
            **common,
            "V04源工作表": rule["source_sheet"], "V04源单元格": rule["source_cell"],
            "V04评分标准原文": rule["standard"], "V04计算公式原文": rule["formula"],
            "Target记录ID": target.get("_record_id", ""), "目标ID": text(target.get("Target_ID")),
            "目标值": display_number(target_value), "Actual记录ID": actual.get("_record_id", ""),
            "实际ID": text(actual.get("Actual_ID")), "实际值": display_number(actual.get("Actual_Value")),
            "达成率": display_percent(rate), "自动得分": display_number(result.get("Auto_Score")),
            "人工分": display_number(result.get("Manual_Score")), "最终得分": display_number(result.get("Final_Score")),
            "权重": display_percent(result.get("Weight")), "加权得分": display_number(result.get("Weighted_Score")),
            "月度总分": display_number(result.get("Monthly_Total")), "档位": display_number(result.get("Commission_Tier_Level")),
            "提成基数": display_number(result.get("Commission_Base")), "提成比例": display_percent(result.get("Commission_Ratio")),
            "提成金额": display_number(result.get("Commission_Amount")), "豁免": text(result.get("Is_Exempt")),
            "结果来源批次": text(result.get("Source")), "复核状态": text(result.get("Review_Status")),
            **d013_review,
        }
        # D-009.3：客服奖惩/定性指标缺少负责人分数时，说明书明确待人工分，绝不把 Base 的公式占位 0 当作业务得分。
        if text(metric.get("Scoring_Type")) in {"定性等级型", "奖惩制"} and result.get("Manual_Score") in (None, ""):
            output_row["人工分"] = "待人工分"
            output_row["最终得分"] = "待人工分"
            output_row["加权得分"] = "待人工分"
        output_row["计算结果与文本对应说明"] = explanation(output_row)
        input_rows.append(input_row)
        output_rows.append(output_row)
    return input_rows, output_rows


def set_header(ws, columns: list[str], fill: PatternFill, font: Font) -> None:
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(2, ws.max_row)}"


def write_table(ws, rows: list[dict[str, Any]], columns: list[str], widths: dict[str, float]) -> None:
    for row in rows:
        ws.append([row.get(c, "") for c in columns])
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    for index, name in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(index)].width = widths.get(name, 14)
    ws.row_dimensions[1].height = 32


def make_summary(output_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[str]]:
    grouped: dict[tuple[str, str, str, str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in output_rows:
        grouped[(row["期间"], row["姓名"], row["部门"], row["岗位"], row["渠道"], row["员工ID"])].append(row)
    metric_names = sorted({r["指标名称"] for r in output_rows})
    summary: list[dict[str, Any]] = []
    for key, rows in sorted(grouped.items()):
        period, name, department, position, channel, employee_id = key
        item = {"期间": period, "姓名": name, "部门": department, "岗位": position, "渠道": channel, "员工ID": employee_id, "指标数": len(rows)}
        for metric in metric_names:
            match = next((r for r in rows if r["指标名称"] == metric), None)
            item[f"【录入】{metric}-目标值"] = match["目标值"] if match else ""
            item[f"【录入】{metric}-实际值"] = match["实际值"] if match else ""
            item[f"【录入】{metric}-达成率"] = match["达成率"] if match else ""
            item[f"【输出】{metric}-得分"] = match["最终得分"] if match else ""
        first = rows[0]
        item.update({"【输出】月度总分": first["月度总分"], "【输出】档位": first["档位"], "【输出】提成金额合计": round(sum(numeric(r["提成金额"]) or 0 for r in rows), 4)})
        summary.append(item)
    columns = ["期间", "姓名", "部门", "岗位", "渠道", "员工ID", "指标数"]
    for metric in metric_names:
        columns += [f"【录入】{metric}-目标值", f"【录入】{metric}-实际值", f"【录入】{metric}-达成率", f"【输出】{metric}-得分"]
    columns += ["【输出】月度总分", "【输出】档位", "【输出】提成金额合计"]
    return summary, columns


def build_workbook(input_rows: list[dict[str, Any]], output_rows: list[dict[str, Any]], source: dict[str, Any]) -> None:
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Arial", color="FFFFFF", bold=True)
    note_fill = PatternFill("solid", fgColor="FFF2CC")

    summary, columns1 = make_summary(output_rows)
    by_department: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in summary:
        by_department[row["部门"]].append(row)
    # 部门名是业务主数据，按原值作为 Sheet 名。无效名称显式失败，禁止静默改名。
    invalid_sheet_chars = set("[]:*?/\\")
    department_names = sorted(by_department)
    for index, department in enumerate(department_names):
        if not department or len(department) > 31 or any(char in invalid_sheet_chars for char in department):
            raise RuntimeError(f"部门名不符合 Excel Sheet 约束，拒绝静默改名：{department!r}")
        ws = wb.active if index == 0 else wb.create_sheet()
        ws.title = department
        set_header(ws, columns1, header_fill, header_font)
        write_table(ws, by_department[department], columns1, {c: 16 for c in columns1} | {"姓名": 12, "部门": 15, "岗位": 18, "渠道": 20, "员工ID": 16})
        ws["A1"].comment = Comment("D-011/D-012：同一员工行内展示负责渠道、录入侧目标/实际/达成率与输出侧得分/总分/档位/提成；明细可按 Result记录ID 下钻。", "data-engineer")

    ws2 = wb.create_sheet("录入明细")
    columns2 = ["期间", "姓名", "部门", "岗位", "渠道", "员工ID", "指标名称", "指标ID", "Target记录ID", "目标ID", "目标值", "Target来源批次", "Actual记录ID", "实际ID", "实际值", "Actual来源批次", "达成率", "Result记录ID", "结果ID", "数据完整性"]
    set_header(ws2, columns2, header_fill, header_font)
    write_table(ws2, input_rows, columns2, {"指标名称": 25, "Target来源批次": 22, "Actual来源批次": 22, "数据完整性": 22})

    ws3 = wb.create_sheet("输出明细（文本+计算对照）")
    columns3 = ["期间", "姓名", "部门", "岗位", "渠道", "员工ID", "指标名称", "指标ID", "评分类型", "V04源工作表", "V04源单元格", "V04评分标准原文", "V04计算公式原文", "Target记录ID", "目标ID", "目标值", "Actual记录ID", "实际ID", "实际值", "达成率", "自动得分", "人工分", "最终得分", "权重", "加权得分", "月度总分", "档位", "提成基数", "提成比例", "提成金额", "绩效工资基数", "档位百分比", "绩效工资部分金额", "中控人数", "渠道GSV", "GSV提成部分金额", "复合收入（两部分合计）", "豁免", "结果来源批次", "复核状态", "计算结果与文本对应说明", "Result记录ID", "结果ID"]
    set_header(ws3, columns3, header_fill, header_font)
    widths3 = {"指标名称": 24, "V04评分标准原文": 55, "V04计算公式原文": 38, "计算结果与文本对应说明": 48, "结果来源批次": 22, "复核状态": 16, "复合收入（两部分合计）": 20, "绩效工资部分金额": 20, "GSV提成部分金额": 20}
    write_table(ws3, output_rows, columns3, widths3)
    for cell in ws3[1]:
        if cell.value in {"V04评分标准原文", "计算结果与文本对应说明"}:
            cell.fill = note_fill
            cell.font = Font(name="Arial", color="000000", bold=True)
    ws3["K1"].comment = Comment(f"权威来源：{source['file']}；SHA-256={source['sha256']}。文本由本地 V04 结构化规则读取。", "data-engineer")

    # 固定文档属性；重跑会幂等覆盖相同业务内容（XLSX ZIP 容器时间戳不作为业务数据）。
    reproducible_time = datetime(2026, 8, 18, tzinfo=timezone.utc)
    wb.properties.created = reproducible_time
    wb.properties.modified = reproducible_time
    wb.properties.title = "绩效结果说明书_20260818_V03"
    wb.properties.subject = "D-011/D-012 按部门、含渠道的录入-输出对照；V04 文本与计算结果可追溯核验"
    wb.save(OUT)


def verify_workbook(expected_rows: int) -> dict[str, Any]:
    wb = load_workbook(OUT, read_only=True, data_only=False)
    expected_tail = ["录入明细", "输出明细（文本+计算对照）"]
    if len(wb.sheetnames) < 3 or wb.sheetnames[-2:] != expected_tail:
        raise RuntimeError(f"Sheet 不符合约定：{wb.sheetnames}")
    input_count = wb["录入明细"].max_row - 1
    output_sheet = wb["输出明细（文本+计算对照）"]
    output_count = output_sheet.max_row - 1
    headers = [cell.value for cell in next(output_sheet.iter_rows(min_row=1, max_row=1))]
    standard_col = headers.index("V04评分标准原文") + 1
    blank_standards = sum(1 for row in output_sheet.iter_rows(min_row=2, values_only=True) if not row[standard_col - 1])
    if input_count != expected_rows or output_count != expected_rows:
        raise RuntimeError(f"行数不一致：录入={input_count} 输出={output_count} Base结果={expected_rows}")
    if blank_standards:
        raise RuntimeError(f"输出明细存在 {blank_standards} 行缺少 V04 评分标准原文")
    d013_headers = ["绩效工资基数", "档位百分比", "绩效工资部分金额", "中控人数", "渠道GSV", "GSV提成部分金额", "复合收入（两部分合计）"]
    missing_d013_headers = [header for header in d013_headers if header not in headers]
    if missing_d013_headers:
        raise RuntimeError(f"输出明细缺少 D-013 人工复核字段：{missing_d013_headers}")
    source_index = headers.index("结果来源批次")
    d013_indexes = [headers.index(header) for header in d013_headers]
    d013_salary_index = headers.index("绩效工资基数")
    d013_rows = [row for row in output_sheet.iter_rows(min_row=2, values_only=True) if row[source_index] == "SIMULATED_T21_D013_D014" and row[d013_salary_index] not in (None, "")]
    if not d013_rows:
        raise RuntimeError("未找到可复核的 D-013 中控收入行")
    d013_missing_values = sum(1 for row in d013_rows if any(row[index] in (None, "") for index in d013_indexes))
    if d013_missing_values:
        raise RuntimeError(f"D-013 行存在 {d013_missing_values} 行缺少人工复核值")
    department_sheets = wb.sheetnames[:-2]
    for sheet_name in department_sheets:
        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if "渠道" not in headers or ws.max_row < 2:
            raise RuntimeError(f"部门 Sheet {sheet_name} 缺少渠道列或员工行")
    return {"sheets": wb.sheetnames, "department_sheets": department_sheets, "input_rows": input_count, "output_rows": output_count, "blank_v04_standard_rows": blank_standards}


def write_exception_log(output_rows: list[dict[str, Any]]) -> dict[str, int]:
    """落盘可复查异常清单；只记录，不修改业务数据或结果。"""
    pending_manual = [
        {"result_id": r["结果ID"], "result_record_id": r["Result记录ID"], "employee_id": r["员工ID"], "metric_id": r["指标ID"], "metric_name": r["指标名称"], "reason": "D-009.3：缺少负责人 Manual_Score，V02 显示待人工分"}
        for r in output_rows if r["最终得分"] == "待人工分"
    ]
    missing_links = [
        {"result_id": r["结果ID"], "result_record_id": r["Result记录ID"], "employee_id": r["员工ID"], "metric_id": r["指标ID"], "target_id": r["目标ID"], "actual_id": r["实际ID"], "reason": "Target 或 Actual 关联缺失/不适用；保留空值，不补造数据"}
        for r in output_rows if not r["目标ID"] or not r["实际ID"]
    ]
    missing_channels = sorted({
        (r["员工ID"], r["姓名"], r["部门"], r["岗位"])
        for r in output_rows if r["渠道"] == "—"
    })
    missing_channel_employees = [
        {"employee_id": employee_id, "employee_name": name, "department": department, "position": position,
         "reason": "Employee.Responsible_Channel_IDs 为空或未能关联到 Channel；V03 按约定显示“—”，不猜测或补造渠道关系"}
        for employee_id, name, department, position in missing_channels
    ]
    payload = {
        "task": "T22", "generated_at_utc": now_utc(), "mode": "READ_ONLY_EXCEPTION_LOG",
        "summary": {"pending_manual_score_rows": len(pending_manual), "missing_target_or_actual_link_rows": len(missing_links), "missing_responsible_channel_employees": len(missing_channel_employees)},
        "pending_manual_score_rows": pending_manual, "missing_target_or_actual_link_rows": missing_links,
        "missing_responsible_channel_employees": missing_channel_employees,
    }
    EXCEPTION_LOG.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return payload["summary"]


def main() -> None:
    global OUT
    parser = argparse.ArgumentParser(description="只读 Base 生成 V03 绩效结果说明书。")
    parser.add_argument("--base-token", default=DEFAULT_BASE_TOKEN, help="Base token；默认项目正式 Base。")
    parser.add_argument("--output", default=str(OUT), help="输出路径；默认写入正式 V03 文件。")
    args = parser.parse_args()
    OUT = Path(args.output).resolve()
    if not Path(CLI).exists():
        raise RuntimeError(f"lark-cli 不可用：{CLI}")
    if not V04_RULES.exists():
        raise RuntimeError(f"缺少 V04 结构化规则：{V04_RULES}")
    v04, source = v04_index()
    data = fetch_data(args.base_token)
    input_rows, output_rows = make_rows(data, v04)
    if not output_rows:
        raise RuntimeError("Performance_Result 没有记录，拒绝生成空结果书。")
    build_workbook(input_rows, output_rows, source)
    exceptions = write_exception_log(output_rows)
    verification = verify_workbook(len(data["results"]))
    print(json.dumps({
        "status": "PASS", "mode": "READ_ONLY_BASE_TO_LOCAL_XLSX", "output": str(OUT.relative_to(ROOT)),
        "exception_log": str(EXCEPTION_LOG.relative_to(ROOT)), "exceptions": exceptions,
        "base_result_rows": len(data["results"]), "v04_source_sha256": source["sha256"], "verification": verification,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        ERROR_LOG.write_text(json.dumps({"task": "T22", "generated_at_utc": now_utc(), "status": "FAILED", "error": str(exc)}, ensure_ascii=False, indent=2), encoding="utf-8")
        print(json.dumps({"status": "FAILED", "error_log": str(ERROR_LOG.relative_to(ROOT)), "error": str(exc)}, ensure_ascii=False), file=sys.stderr)
        sys.exit(2)
