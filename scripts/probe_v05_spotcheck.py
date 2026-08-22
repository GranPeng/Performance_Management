#!/usr/bin/env python3
"""补充核查：V04 编导/视频剪辑的爆款激励块位置、V04 视频剪辑 B21、提成基数单元格（只读）。"""
from openpyxl import load_workbook
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
wb4 = load_workbook(ROOT.parent / "1.中期预算调整项目/3.outputs/3.岗位绩效/岗位绩效说明书汇总_20260809_V04_最新绩效框架.xlsx", data_only=False)
wb5 = load_workbook(ROOT / "1.reference/绩效模板/岗位绩效说明书汇总_20260821_V05.xlsx", data_only=False)

for tag, wb in [("V04", wb4), ("V05", wb5)]:
    for name in ["制作部-编导", "制作部-视频剪辑"]:
        ws = wb[name]
        hits = []
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None and ("爆款" in str(c.value) or str(c.value) == "、"):
                    hits.append(f"{c.coordinate}={str(c.value)[:60]}")
        print(f"{tag} {name}: 爆款/顿号相关 -> {hits}")
    print(f"{tag} 视频剪辑 B21 =", repr(wb["制作部-视频剪辑"]["B21"].value))

print("V04 运营 K12 (提成基数) =", repr(wb4["兴趣电商-运营"]["K12"].value))
print("V04 高级主管 M12 (提成基数) =", repr(wb4["兴趣电商-高级主管"]["M12"].value))
ws5 = wb5["提成比例与绩效工资核算规则"]
print("V05 规则表 F32 (高级主管基数) =", repr(ws5["F32"].value), "| F39 (运营基数) =", repr(ws5["F39"].value))
print("V05 规则表 J10/J11 =", repr(ws5["J10"].value), "|", repr(ws5["J11"].value))
print("V05 编导 F16:G19 =", [(c, repr(wb5["制作部-编导"][c].value)) for c in ["F16","F17","G17","F18","G18","F19","G19"]])
