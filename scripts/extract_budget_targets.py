#!/usr/bin/env python3
"""V60「总表-渠道二维汇总」只读提取与KPI映射。

输入：1.reference/羊喜市中期预算模型_20260716_V60.xlsx、data/output/绩效框架结构化规则.json
输出：预算目标结构化提取 JSON、两份 Markdown 文档、校验日志、数据处理记忆。
不修改任何源工作簿。运行前后均记录 SHA-256，用于可追溯和回滚。
"""
from __future__ import annotations

import hashlib
import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "1.reference/羊喜市中期预算模型_20260716_V60.xlsx"
KPI_SOURCE = ROOT / "data/output/绩效框架结构化规则.json"
OUT_JSON = ROOT / "data/output/预算目标结构化提取.json"
OUT_DOC = ROOT / "docs/预算目标结构化提取.md"
OUT_MAP = ROOT / "docs/指标映射表.md"
OUT_LOG = ROOT / "data/output/预算目标提取校验日志.json"
OUT_MEMORY = ROOT / ".hermes/memory/data_memory.md"
SHEET = "总表-渠道二维汇总"


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def native(value: Any) -> Any:
    """保留数值和公式错误；将 openpyxl 不能 JSON 序列化的值转换为文本。"""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def line_label(ws, row: int) -> tuple[str, str | None, str | None]:
    """按照 B/C/D/E 层级返回展示标签、一级分类、二级分类。"""
    b, c, d, e = [ws.cell(row, col).value for col in range(2, 6)]
    if b:
        return str(b), str(b), None
    if c:
        return str(c), str(c), None
    if d and e:
        return f"{d}/{e}", str(d), str(e)
    if d:
        return str(d), str(d), None
    if e:
        return str(e), None, str(e)
    return f"未命名行{row}", None, None


def source_ref(channel: str, cell: str, formula: Any, cached: Any) -> dict[str, Any]:
    return {
        "source_sheet": SHEET,
        "source_channel": channel,
        "source_cell": cell,
        "formula": native(formula),
        "cached_value": native(cached),
    }


def mapping_for(metric: dict[str, Any]) -> dict[str, Any]:
    name = metric["name"]
    base = {
        "source_sheet": metric["source_sheet"],
        "source_cell": metric["source_cell"],
        "position": metric["position"],
        "kpi_name": name,
        "kpi_formula": metric.get("calculation_formula"),
        "kpi_target_definition": metric.get("target"),
    }
    if name in {"团队GSV营收目标达成率", "团队营收目标达成率"}:
        return base | {
            "mapping_status": "候选映射，需业务确认",
            "budget_field": "主营业务收入",
            "budget_field_row": 3,
            "channel_scope": "V60全部9个渠道；具体岗位的负责渠道未在T1结构化清单中给出，不能强行归属",
            "target_value_definition": "按渠道×自然月读取，单位万元；KPI要求“退货后营收（GSV）”或“营收”，V60字段仅标注“主营业务收入”。",
            "reason": "名称均指收入/营收目标，但V60未标明是否为退货后GSV，缺少口径确认。",
        }
    if name in {"团队ROI签收目标达成率", "个人直播签收ROI达成"}:
        return base | {
            "mapping_status": "未映射",
            "budget_field": None,
            "budget_field_row": None,
            "channel_scope": "不适用",
            "target_value_definition": None,
            "reason": "V60目标表无“签收ROI”字段；“广告投流费”“销售费用率”等均不能证明为签收ROI分母/分子或签收口径。",
        }
    if name == "团队产出数量":
        return base | {
            "mapping_status": "未映射",
            "budget_field": None,
            "budget_field_row": None,
            "channel_scope": "不适用",
            "target_value_definition": None,
            "reason": "V60仅含金额型损益与费用科目，无剪辑团队人数、视频产出数量或人均产出字段。",
        }
    if name == "团队产出消耗金额":
        return base | {
            "mapping_status": "口径不一致，不能直接映射",
            "budget_field": "广告投流费（候选）",
            "budget_field_row": 11,
            "channel_scope": "V60全部9个渠道；未细分项目组/素材上传90天窗口",
            "target_value_definition": "按渠道×自然月读取，单位万元。",
            "reason": "KPI要求项目组90天内成片对应的总消耗与项目预算目标；V60为渠道损益中的广告投流费，缺少项目组、素材归因及90天窗口。",
        }
    if name in {"个人成片数量", "个人素材数量"}:
        return base | {
            "mapping_status": "未映射",
            "budget_field": None,
            "budget_field_row": None,
            "channel_scope": "不适用",
            "target_value_definition": None,
            "reason": "V60不含个人成片、素材数量、工单任务数等生产任务数据。",
        }
    if name == "个人成片消耗金额":
        return base | {
            "mapping_status": "口径不一致，不能直接映射",
            "budget_field": "广告投流费（候选）",
            "budget_field_row": 11,
            "channel_scope": "V60全部9个渠道；未细分项目/个人/剪辑人数",
            "target_value_definition": "按渠道×自然月读取，单位万元。",
            "reason": "KPI需要个人消耗、项目总消耗、项目组剪辑人数及90天素材范围；V60只有渠道级广告投流费。",
        }
    raise ValueError(f"未覆盖的定量KPI：{name}")


def md_table(headers: list[str], rows: list[list[Any]]) -> str:
    def clean(v: Any) -> str:
        return str(v if v is not None else "—").replace("|", "\\|").replace("\n", "<br>")
    return "| " + " | ".join(headers) + " |\n| " + " | ".join(["---"] * len(headers)) + " |\n" + "\n".join(
        "| " + " | ".join(clean(v) for v in row) + " |" for row in rows
    )


def main() -> None:
    for p in [SOURCE, KPI_SOURCE]:
        if not p.exists():
            raise FileNotFoundError(p)
    source_hash_before = sha256(SOURCE)
    kpi = json.loads(KPI_SOURCE.read_text(encoding="utf-8"))
    # 使用普通加载以避免 read_only 模式对随机单元格访问的重复全表扫描；全程不调用 save()，源文件仍为只读提取。
    wb_formula = load_workbook(SOURCE, read_only=False, data_only=False)
    wb_value = load_workbook(SOURCE, read_only=False, data_only=True)
    ws_f = wb_formula[SHEET]
    ws_v = wb_value[SHEET]

    blocks = []
    for col in range(1, ws_f.max_column + 1):
        header = ws_f.cell(1, col).value
        if isinstance(header, str) and header.endswith("（万元）") and header != "渠道总贡献（万元）":
            channel = header.removesuffix("（万元）")
            months = []
            for month_col in range(col, col + 12):
                month = ws_f.cell(2, month_col).value
                if not isinstance(month, str) or not month.startswith("2026-"):
                    raise ValueError(f"{channel} {get_column_letter(month_col)}2 月份表头异常：{month!r}")
                months.append((month, month_col))
            blocks.append({"channel": channel, "start_column": col, "months": months})
    if len(blocks) != 9:
        raise ValueError(f"渠道块数量应为9，实际为{len(blocks)}")

    all_rows = []
    errors = []
    for row in range(3, ws_f.max_row + 1):
        label, category, subcategory = line_label(ws_f, row)
        # 仅跳过空行与解释性分析行，保留所有预算科目、合计和比率行。
        if row in (43, 61, 62) or label.startswith("未命名行"):
            continue
        item = {
            "source_row": row,
            "line_item": label,
            "category": category,
            "subcategory": subcategory,
            "unit": "万元" if "率" not in label and label != "Margin" else "比例",
            "channels": [],
        }
        for block in blocks:
            values = []
            for month, col in block["months"]:
                formula_cell = ws_f.cell(row, col)
                value_cell = ws_v.cell(row, col)
                raw = native(value_cell.value)
                formula = native(formula_cell.value)
                if raw == "#REF!" or (isinstance(formula, str) and "#REF!" in formula):
                    errors.append({
                        "severity": "warning",
                        "code": "SOURCE_FORMULA_REF_ERROR",
                        "source_sheet": SHEET,
                        "source_cell": formula_cell.coordinate,
                        "line_item": label,
                        "channel": block["channel"],
                        "month": month,
                        "message": "源工作簿已存在 #REF!；原样记录，未修改。",
                    })
                values.append({
                    "month": month,
                    "value": raw,
                    "source": source_ref(block["channel"], f"{get_column_letter(col)}{row}", formula, raw),
                })
            item["channels"].append({"channel": block["channel"], "monthly_values": values})
        all_rows.append(item)

    # 扫描整个目标工作表的公式错误，而不仅限于渠道月度值；异常仅记录、不修复。
    recorded_error_cells = {e["source_cell"] for e in errors}
    for row in ws_f.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "#REF!" in cell.value and cell.coordinate not in recorded_error_cells:
                errors.append({
                    "severity": "warning",
                    "code": "SOURCE_FORMULA_REF_ERROR",
                    "source_sheet": SHEET,
                    "source_cell": cell.coordinate,
                    "line_item": line_label(ws_f, cell.row)[0],
                    "channel": None,
                    "month": None,
                    "message": "源工作簿已存在 #REF!；原样记录，未修改。",
                })
                recorded_error_cells.add(cell.coordinate)

    quantitative = []
    for pos in kpi["positions"]:
        for metric in pos["metrics"]:
            if metric.get("scoring_type") == "达成率/进度型":
                quantitative.append({
                    "source_sheet": pos["source_sheet"],
                    "position": pos["position"],
                    "source_cell": metric["source_cell"],
                    "name": metric["name"],
                    "calculation_formula": metric.get("calculation_formula"),
                    "target": metric.get("target"),
                })
    mappings = [mapping_for(m) for m in quantitative]
    status_counts = Counter(m["mapping_status"] for m in mappings)

    output = {
        "schema_version": "1.0",
        "record_type": "budget_target_extraction_and_kpi_mapping",
        "source": {
            "file": SOURCE.name,
            "sha256": source_hash_before,
            "source_sheet": SHEET,
            "read_mode": "openpyxl data_only=False/True 双读取；全程未调用 save()，只读提取，未修改源文件",
            "unit_note": "工作表渠道标题标注为（万元）；比例行按原单元格值保留。",
        },
        "extraction": {
            "timestamp_utc": datetime.now(timezone.utc).isoformat(),
            "channels": [b["channel"] for b in blocks],
            "months": [m for m, _ in blocks[0]["months"]],
            "line_item_count": len(all_rows),
            "channel_month_value_count": len(all_rows) * len(blocks) * 12,
            "transform": "按第1行渠道块、第2行月份表头、第3-60行预算项目抽取；每个值保留源单元格、公式和缓存值。",
            "source_formula_error_policy": "发现源 #REF! 时不修正，输出为原始异常并写入校验日志。",
        },
        "budget_targets": all_rows,
        "kpi_mapping": {
            "anchor_file": KPI_SOURCE.name,
            "quantitative_metric_count": len(mappings),
            "status_counts": dict(status_counts),
            "mappings": mappings,
        },
    }
    source_hash_after = sha256(SOURCE)
    validation = {
        "schema_version": "1.0",
        "record_type": "budget_target_extraction_validation_log",
        "run_timestamp_utc": datetime.now(timezone.utc).isoformat(),
        "source_sha256_before": source_hash_before,
        "source_sha256_after": source_hash_after,
        "checks": [
            {"name": "source_unchanged", "passed": source_hash_before == source_hash_after, "expected": source_hash_before, "actual": source_hash_after},
            {"name": "channel_count", "passed": len(blocks) == 9, "expected": 9, "actual": len(blocks)},
            {"name": "month_count_per_channel", "passed": all(len(x["months"]) == 12 for x in blocks), "expected": 12, "actual": [len(x["months"]) for x in blocks]},
            {"name": "quantitative_kpi_mapping_coverage", "passed": len(mappings) == len(quantitative), "expected": len(quantitative), "actual": len(mappings)},
            {"name": "source_reference_completeness", "passed": all(v["source"]["source_cell"] for i in all_rows for c in i["channels"] for v in c["monthly_values"]), "expected": "all extracted values carry source cells", "actual": "verified"},
        ],
        "exceptions": errors,
        "exception_summary": dict(Counter(e["code"] for e in errors)),
    }
    if not all(c["passed"] for c in validation["checks"]):
        raise RuntimeError("结构校验失败，拒绝写入产出")

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_DOC.parent.mkdir(parents=True, exist_ok=True)
    OUT_MEMORY.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(output, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    OUT_LOG.write_text(json.dumps(validation, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    doc_rows = []
    for item in all_rows:
        if item["source_row"] in (3, 8, 11, 41, 56, 59):
            for channel in item["channels"]:
                jan = channel["monthly_values"][0]
                jun = channel["monthly_values"][5]
                dec = channel["monthly_values"][11]
                doc_rows.append([item["source_row"], item["line_item"], channel["channel"], jan["value"], jun["value"], dec["value"], jan["source"]["source_cell"]])
    OUT_DOC.write_text(
        "# V60 分渠道分月预算目标结构化提取\n\n"
        "## 提取范围与可追溯性\n"
        f"- 权威源：`{SOURCE.name}` / `{SHEET}`（只读，SHA-256：`{source_hash_before}`）。\n"
        "- 颗粒度：渠道 × 月份 × 预算项目；9 个渠道、12 个自然月、"
        f"{len(all_rows)} 个预算项目，共 {len(all_rows) * len(blocks) * 12} 个渠道月度值。\n"
        "- 单位：渠道标题标注为万元；比率/Margin类按源单元格比例值保留。\n"
        "- 完整机器可读明细：`data/output/预算目标结构化提取.json`；每个值均含源单元格、原公式、缓存值。\n"
        "- 源文件异常：发现 `#REF!` 时仅记录，不作修复或业务推断；详见校验日志。\n\n"
        "## 渠道与月份\n"
        f"- 渠道：{'、'.join(b['channel'] for b in blocks)}\n"
        f"- 月份：{'、'.join(m for m, _ in blocks[0]['months'])}\n\n"
        "## 关键预算字段抽查（源缓存值）\n"
        + md_table(["源行", "预算字段", "渠道", "2026-01", "2026-06", "2026-12", "1月源单元格"], doc_rows)
        + "\n\n## 完整预算字段清单\n"
        + md_table(["源行", "预算字段", "一级分类", "二级分类", "单位"], [[x["source_row"], x["line_item"], x["category"], x["subcategory"], x["unit"]] for x in all_rows])
        + "\n\n## 校验结论\n"
        f"- 9 个渠道均有连续12个月份表头；定量KPI {len(mappings)} 项均已进入映射表。\n"
        f"- 源文件 SHA-256 运行前后一致：`{source_hash_after}`。\n"
        f"- 检出源公式 `#REF!` {len(errors)} 个单元格，全部原样保留并记录，未修改源工作簿。\n",
        encoding="utf-8",
    )

    map_rows = [[m["position"], m["kpi_name"], m["source_sheet"], m["source_cell"], m["mapping_status"], m["budget_field"], m["budget_field_row"], m["channel_scope"], m["target_value_definition"], m["reason"]] for m in mappings]
    unmapped = [m for m in mappings if m["mapping_status"] != "候选映射，需业务确认"]
    OUT_MAP.write_text(
        "# 绩效指标 ↔ V60预算字段映射表\n\n"
        "## 映射原则\n"
        "- 以 T1 的 `data/output/绩效框架结构化规则.json` 中“达成率/进度型”指标为完整锚点；不改变KPI或预算字段的业务含义。\n"
        "- “候选映射”不等于可直接用于计算：若源名称、退货/签收范围、归因颗粒度或时间窗口未被源文件明确支持，必须保留为待确认。\n"
        "- V60源字段的统一定位：`总表-渠道二维汇总`，渠道块第1行、月份第2行；“主营业务收入”为第3行，“广告投流费”为第11行。\n\n"
        "## 全量定量KPI映射\n"
        + md_table(["岗位", "KPI", "T1来源", "单元格", "映射状态", "V60字段", "V60行", "渠道归属", "目标值口径", "备注/原因"], map_rows)
        + "\n\n## 未能直接对齐的字段及原因\n"
        + md_table(["KPI", "状态", "候选/缺失字段", "原因"], [[m["kpi_name"], m["mapping_status"], m["budget_field"], m["reason"]] for m in unmapped])
        + "\n\n## 待业务确认事项\n"
        "1. V60“主营业务收入”是否等同于绩效中的“退货后营收（GSV）”；如不等同，需提供退货后GSV的权威预算来源与渠道映射。\n"
        "2. 签收ROI的预算目标在何处维护；不得以广告投流费、销售费用率等字段自行反推。\n"
        "3. V60渠道与岗位“所负责渠道”的关系表，以及项目组/个人/素材90天归因的数据来源。\n",
        encoding="utf-8",
    )
    OUT_MEMORY.write_text(
        "# 数据处理记忆\n\n"
        "## DM-001：V60预算目标提取与绩效映射口径发现\n"
        f"- 日期：{datetime.now().date().isoformat()}\n"
        f"- 权威源：`{SOURCE.name}` 的 `{SHEET}`；读取范围为9渠道×12月份，渠道标题单位为万元。\n"
        "- `主营业务收入`（源行3）可作为GSV/营收类KPI的候选预算字段，但V60未明确“退货后GSV”口径，禁止未经确认直接作为绩效分母。\n"
        "- V60未包含签收ROI预算、个人/团队产出数量、工单、素材数量、项目组人数及90天归因数据；这些KPI需要补充权威数据源。\n"
        "- `广告投流费`（源行11）可供“消耗金额”相关分析参考，但因缺项目组/个人/90天窗口，不能直接作为该类KPI预算目标。\n"
        f"- 源表的退货成本汇总公式存在 {len(errors)} 个 `#REF!` 渠道月度异常；本任务只记录，不修改源文件。\n",
        encoding="utf-8",
    )
    print(json.dumps({
        "status": "success",
        "source_sha256": source_hash_before,
        "channels": len(blocks),
        "months": 12,
        "line_items": len(all_rows),
        "channel_month_values": len(all_rows) * len(blocks) * 12,
        "quantitative_kpis": len(mappings),
        "mapping_status_counts": dict(status_counts),
        "source_exceptions": len(errors),
        "outputs": [str(x.relative_to(ROOT)) for x in [OUT_JSON, OUT_DOC, OUT_MAP, OUT_LOG, OUT_MEMORY]],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
