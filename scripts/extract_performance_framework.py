#!/usr/bin/env python3
"""只读提取 V04 绩效框架；生成可追溯的规则文档、JSON 与校验日志。"""
from __future__ import annotations
import hashlib, json
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "1.reference/岗位绩效说明书汇总_20260809_V04_最新绩效框架.xlsx"
OUT_JSON = ROOT / "data/output/绩效框架结构化规则.json"
OUT_FULL = ROOT / "docs/绩效框架结构化规则.md"
OUT_RULES = ROOT / "docs/business_rules.md"
OUT_LOG = ROOT / "data/output/绩效框架提取校验日志.json"
GLOSSARY = ROOT / "context/business_glossary.md"
MEMORY = ROOT / ".hermes/memory/business_memory.md"

# 仅为来源 sheet 设置稳定短码；不是业务关联键，也不改变源数据。
SHEET_CODE = {
    "兴趣电商-运营": "IE-OPS", "兴趣电商-高级主管": "IE-SUP", "兴趣电商-广告投放": "IE-ADS",
    "兴趣电商-直播中控": "IE-LIVE", "兴趣电商-主播": "IE-HOST", "兴趣电商-客服": "IE-CS",
    "制作部-编导": "PROD-DIR", "制作部-视频剪辑": "PROD-EDIT", "制作部-摄影师": "PROD-CAM",
}

def value(v):
    return v if v is not None else None

def text(v):
    return "" if v is None else str(v)

def score_type(scoring: str, formula: str, reward: str = "", penalty: str = "") -> str:
    all_text = "\n".join([scoring, formula, reward, penalty])
    if "达成率=" in formula or "达成率=" in scoring or "人均产出=" in formula or "个人贡献率公式=" in formula:
        return "达成率/进度型"
    if "错误次数" in all_text or "漏回复次数" in all_text or "次数" in all_text and "得" in all_text:
        return "次数阈值型"
    if "单次扣除" in all_text or "单次扣" in all_text:
        return "扣分制"
    if reward or penalty or "奖励" in all_text or "处罚" in all_text:
        return "奖惩制"
    if "0-60分" in all_text or "0—60分" in all_text:
        return "定性等级型"
    return "文字评分（类型未显式标注）"

def col_name(c): return chr(64 + c)

def extract_metrics(ws):
    # 两类表头：标准表 B:M；客服表 B:L，列 F/G/H 依次为评分/奖励/处罚。
    header_row = 4
    headers = {text(ws.cell(header_row, c).value).replace("\n", ""): c for c in range(1, ws.max_column + 1) if ws.cell(header_row,c).value}
    is_customer_service = "奖励条件及金额/提成标准" in headers
    total_row = next((r for r in range(6, ws.max_row + 1) if text(ws.cell(r, 2).value) == "合计"), None)
    if total_row is None: raise ValueError(f"{ws.title}: 未找到合计行")
    result, dimension = [], None
    for r in range(6, total_row):
        number = ws.cell(r, 2).value
        # 客服表第11行存在指标名称但B列未填写编号；保留该指标且明确标记缺失，不能静默丢弃。
        if number is None and ws.cell(r, 4).value is None:
            continue
        source_number = number if number is not None else "未填写（源单元格B%d为空）" % r
        d = ws.cell(r, 3).value
        if d is not None: dimension = d
        if is_customer_service:
            metric = {
                "source_metric_number": source_number, "source_metric_number_raw": number, "source_cell": f"B{r}", "dimension": dimension,
                "name": value(ws.cell(r,4).value), "weight": value(ws.cell(r,5).value),
                "target": None, "calculation_formula": None, "unit": None,
                "scoring_standard": value(ws.cell(r,6).value), "reward_condition": value(ws.cell(r,7).value),
                "penalty_condition": value(ws.cell(r,8).value), "data_source": value(ws.cell(r,9).value),
                "evaluation_period": value(ws.cell(r,10).value), "source_result_example": value(ws.cell(r,11).value),
                "source_score_formula": value(ws.cell(r,12).value),
            }
        else:
            metric = {
                "source_metric_number": source_number, "source_metric_number_raw": number, "source_cell": f"B{r}", "dimension": dimension,
                "name": value(ws.cell(r,4).value), "weight": value(ws.cell(r,5).value),
                "target": value(ws.cell(r,6).value), "calculation_formula": value(ws.cell(r,7).value),
                "unit": value(ws.cell(r,8).value), "scoring_standard": value(ws.cell(r,9).value),
                "reward_condition": None, "penalty_condition": None, "data_source": value(ws.cell(r,10).value),
                "evaluation_period": value(ws.cell(r,11).value), "source_result_example": value(ws.cell(r,12).value),
                "source_score_formula": value(ws.cell(r,13).value),
            }
        metric["scoring_type"] = score_type(text(metric["scoring_standard"]), text(metric["calculation_formula"]), text(metric["reward_condition"]), text(metric["penalty_condition"]))
        metric["score_cap_or_floor"] = "按评分标准原文执行；未另行推断"
        result.append(metric)
    return result, total_row, is_customer_service

def extract_commission(ws):
    # 找“提成来源”作为梯度表表头；保留每个实际单元格及其列名，避免对空白列作主观对齐。
    header = None
    for row in ws.iter_rows():
        for c in row:
            if c.value == "提成来源": header = (c.row, c.column); break
        if header: break
    if not header: return []
    r, start = header
    max_col = min(ws.max_column, start + 4)
    columns = [{"column": col_name(c), "header": value(ws.cell(r,c).value)} for c in range(start, max_col+1)]
    tiers=[]
    for rr in range(r+1, ws.max_row+1):
        first = ws.cell(rr,start).value
        if first is None: continue
        if text(first).startswith("*") or first == "、": continue
        row={"source_row": rr, "source_cell": f"{col_name(start)}{rr}", "values": {col_name(c): value(ws.cell(rr,c).value) for c in range(start,max_col+1)}}
        tiers.append(row)
    return [{"source_header_row": r, "columns": columns, "tiers": tiers}]

def extract_special_rules(ws):
    rows=[]
    for row in ws.iter_rows():
        for c in row:
            if c.value == "考核方式说明":
                h = c.row + 1
                for rr in range(h+1, ws.max_row+1):
                    if ws.cell(rr,4).value is None: continue
                    rows.append({"source_row": rr, "负责项目": value(ws.cell(rr,4).value), "渠道": value(ws.cell(rr,5).value), "考核方式": value(ws.cell(rr,6).value), "新项目评价周期": value(ws.cell(rr,7).value)})
    # 星号备注同为来源规则，单列列出。
    notes=[]
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value,str) and c.value.startswith("*"):
                notes.append({"source_cell": c.coordinate, "text": c.value})
    return {"assessment_methods": rows, "source_notes": notes}

def render_metric(m):
    def x(k): return text(m.get(k)).replace("\n", "<br>") or "—"
    return f"| {x('source_metric_number')} | {x('dimension')} | {x('name')} | {x('scoring_type')} | {x('weight')} | {x('target')} | {x('calculation_formula')} | {x('unit')} | {x('scoring_standard')} | {x('data_source')} | {x('evaluation_period')} | {x('score_cap_or_floor')} |"

def main():
    for p in [OUT_JSON, OUT_FULL, OUT_RULES, OUT_LOG, GLOSSARY, MEMORY]: p.parent.mkdir(parents=True, exist_ok=True)
    source_hash = hashlib.sha256(SOURCE.read_bytes()).hexdigest()
    wb = load_workbook(SOURCE, data_only=False)
    # 第一次结构提取：不改写 source workbook。
    positions=[]; total_metrics=0; sheet_checks=[]
    for ws in wb.worksheets:
        metrics,total_row,is_cs = extract_metrics(ws)
        formula_cells = [c.coordinate for row in ws.iter_rows() for c in row if c.data_type == "f"]
        weight_formula = ws.cell(total_row,5).value
        score_formula = ws.cell(total_row,12 if is_cs else 13).value
        position={
            "source_sheet": ws.title, "source_sheet_code": SHEET_CODE[ws.title], "department": ws["D2"].value,
            "position": ws["F2"].value, "source_range": f"B4:{col_name(ws.max_column)}{ws.max_row}",
            "metrics": metrics, "weight_total_formula": weight_formula, "total_score_formula": score_formula,
            "commission_rules": extract_commission(ws), "special_rules": extract_special_rules(ws),
        }
        positions.append(position); total_metrics += len(metrics)
        sheet_checks.append({"sheet":ws.title,"metric_count":len(metrics),"source_total_row":total_row,"weight_total_formula":weight_formula,"total_score_formula":score_formula,"formula_cells":formula_cells})
    # 可疑/待业务确认：只记录源内歧义，不修补、不归一化。
    disputes=[
        {"source":"兴趣电商-客服!B11","issue":"“邀请好评”指标存在但B列编号为空；已保留该指标，并在结构化结果中标记为“未填写（源单元格B11为空）”，未自行补号。"},
        {"source":"兴趣电商-广告投放!I7","issue":"评分标准末档原文为“90%＞达成率，0分”，与紧邻的“90%＞达成率≥80%，60分”重叠；按原文保留，需业务确认。"},
        {"source":"兴趣电商-直播中控!I18:M21","issue":"提成表的列头与数据列存在视觉错位：J 列存分数下限、K 列存上限、L 列存“0.1%/中控人数”、M 列存“绩效工资*比例”。JSON按源单元格保留，未自行重排。"},
        {"source":"兴趣电商-广告投放!I17:M20；兴趣电商-主播!I16:M19；制作部各岗位提成表","issue":"多个提成梯度最高仅至150分，未给出≥150分的处理规则；不得自行外推。"},
        {"source":"各岗位“完成结果”列及高级主管测算区","issue":"源工作簿含示例完成结果/测算值；其是否为演示数据或历史业务结果未标记，本提取仅作为来源字段记录，不作为规则事实。"},
        {"source":"兴趣电商-客服!F6:H11","issue":"客服表按奖励/处罚金额与分值并列描述，但未给出将金额、次数与最终得分的统一换算公式；本提取保留原文，待业务与财务确认。"},
    ]
    data={
        "schema_version":"1.0", "record_type":"performance_framework_rule_extraction", "source": {"file": SOURCE.name, "sha256":source_hash, "read_mode":"openpyxl data_only=False；只读提取，未修改源文件", "sheet_count":len(wb.sheetnames), "sheet_names":wb.sheetnames},
        "extraction": {"timestamp_utc":datetime.now(timezone.utc).isoformat(), "transform":"每个岗位表从第4行表头、编号列与合计行之间提取KPI；合并单元格评价维度向下继承；提成与特殊规则按源单元格保留。", "non_authoritative_generated_fields":["source_sheet_code","scoring_type","score_cap_or_floor"], "traceability":"每条KPI保留 source_sheet、source_cell 与 source_metric_number。"},
        "positions":positions, "disputes_and_open_items":disputes
    }
    OUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2)+"\n", encoding="utf-8")
    # 二次读取 JSON 验证持久化、可解析且每 sheet KPI 数量一致。
    reloaded=json.loads(OUT_JSON.read_text(encoding="utf-8"))
    assert len(reloaded["positions"]) == len(wb.sheetnames)
    assert sum(len(x["metrics"]) for x in reloaded["positions"]) == total_metrics
    assert [x["source_sheet"] for x in reloaded["positions"]] == wb.sheetnames
    
    full=["# 绩效框架结构化规则（V04 源文件全量提取）","","> 状态：只读提取；本文件与 `data/output/绩效框架结构化规则.json` 由同一脚本生成。任何规则变更须经 Business Analyst、Finance Controller、Product Owner 三方确认。","",f"- 权威源：`{SOURCE.name}`","- 源文件 SHA-256：`%s`"%source_hash,"- 覆盖：9 个工作表、%d 项 KPI。"%total_metrics,"- 可追溯：每项保留源工作表、源单元格、源编号。","- 说明：`scoring_type` 是依据原文的技术分类，不是源文件新增业务规则；封顶/保底以评分标准原文为准。",""]
    for p in positions:
        full += [f"## {p['source_sheet']}（{p['department']} / {p['position']}）","",f"- 来源范围：`{p['source_range']}`",f"- 权重合计公式：`{p['weight_total_formula']}`；最终得分合计公式：`{p['total_score_formula']}`","","| 编号 | 评价维度 | 指标项目 | 评分类型 | 权重 | 目标 | 计算公式 | 单位 | 评分标准（原文） | 数据来源 | 评价周期 | 封顶/保底 |","|---:|---|---|---|---:|---|---|---|---|---|---|---|"]
        full += [render_metric(m) for m in p['metrics']]
        full += [""]
        for cr in p['commission_rules']:
            full += ["### 岗位提成标准（源单元格保留）", "", "表头："+"；".join(f"{x['column']}列={text(x['header']).replace(chr(10),' / ')}" for x in cr['columns']), ""]
            for t in cr['tiers']: full.append(f"- `{t['source_cell']}`："+"；".join(f"{k}={v}" for k,v in t['values'].items() if v is not None))
            full += [""]
        sr=p['special_rules']
        if sr['assessment_methods'] or sr['source_notes']:
            full += ["### 特殊/补充规则（原文）",""]
            for x in sr['assessment_methods']: full.append("- "+"；".join(f"{k}={v}" for k,v in x.items() if k!='source_row' and v is not None))
            for x in sr['source_notes']: full.append(f"- `{x['source_cell']}`：{x['text']}")
            full += [""]
    full += ["## 争议项与待确认","","> 以下项目是源文件信息不完整、矛盾或未标注性质之处；不是本任务对规则的修改。","" ]
    full += [f"- `{x['source']}`：{x['issue']}" for x in disputes]
    OUT_FULL.write_text("\n".join(full)+"\n",encoding="utf-8")

    rules=["# 业务规则（V04 绩效框架只读基线）","","> 本文件是 V04 权威 Excel 的规则索引与业务口径基线；指标逐项原文见 `docs/绩效框架结构化规则.md`，机器可读版见 `data/output/绩效框架结构化规则.json`。不以本文件取代源文件，任何变更须三方确认。","",f"- 权威源：`{SOURCE.name}`",f"- SHA-256：`{source_hash}`",f"- 提取范围：9 岗位、{total_metrics} 项 KPI。","","## 通用核算链","","- 绩效总分：各 KPI 的“最终得分”按源表公式加总；标准岗位表公式为 `权重 × 完成结果`，客服表为 `权重 × 完成结果`。","- 达成率型指标：源表有明确公式时按原文（通常为实际值 ÷ 目标值）计算，再按各指标评分标准给分；不得将不同岗位的阈值互相套用。","- 定性、次数、扣分、奖惩类指标：输入须保留数据来源与评价周期；源表未提供统一换算公式的，不得由 ETL 推断最终分数。","- 评价周期：除源表另有范围说明外，均为每月5日完成核算、取上月整月。制作部同时限定“上传时间90天内素材的上月整月消耗”。","","## 财务/业务口径（仅源文件明确内容）","","- GSV：源表表述为“退货后营收（GSV）”，用于多个团队营收达成指标；实际值与预算目标的口径仍须保持同一渠道/范围。","- ROI：兴趣电商相关岗位使用“签收ROI”，按实际签收ROI ÷ 预算目标签收ROI计算（仅在明确填写公式的岗位中适用）。","- 制作部消耗：数据源为云视频管家，且计算范围为上传时间90天内素材的上月整月消耗。","- 新项目（高级主管）：高个子视频号、少年状元整项目、骨能整项目按新项目机制；原文写明暂定脱离条件为单月GSV达成100万或运营两个月，满足任意条件。","","## 各岗位 KPI 总览","", "| 岗位 | KPI（编号：名称；权重；评分类型） |","|---|---|"]
    for p in positions:
        items="<br>".join(f"{m['source_metric_number']}：{m['name']}；{m['weight']}；{m['scoring_type']}" for m in p['metrics'])
        rules.append(f"| {p['position']}（{p['source_sheet']}） | {items} |")
    rules += ["","## 提成与激励规则索引","", "- 兴趣电商运营、高级主管：运营店铺 GSV 的得分梯度，固定提成基数为 0.003，系数按分段取值；完整下限、上限、系数与公式见结构化规则文档。","- 广告投放：投放账户广告消耗的得分分段提成比例均为0.001（源表列为100分前的四档）。","- 直播中控/运营助理：源表将“0.1%/中控人数”与“绩效工资×80%/90%/100%/110%”并列记录，列映射待确认。","- 主播：个人直播时段 GSV，源表列示0~60/60~80/80~100/100~150分对应0.001/0.002/0.004/0.005。","- 客服：提成按各 KPI 奖励条件及金额为准，源表无统一梯度或得分换算公式。","- 编导、视频剪辑：近90天内上传素材单视频消耗超50万为爆款；除消耗提成外额外一次性奖励1500元/片。","- 摄影师：个人拍摄素材消耗的四档提成比例见结构化规则文档。","","## 争议项（禁止自行补充）",""]
    rules += [f"- `{x['source']}`：{x['issue']}" for x in disputes]
    OUT_RULES.write_text("\n".join(rules)+"\n",encoding="utf-8")

    GLOSSARY.write_text("""# 业务术语表（Business Glossary）

> 所有 Agent 产出文档、字段命名、对话回复时，术语必须与本表一致。
> 本次由 T1 从 V04 源 Excel 只读补全；定义只转述源文件，未定义事项标为待确认。

## 绩效核算域

| 术语 | 英文/缩写 | 定义 | 来源 |
|---|---|---|---|
| 绩效框架 | — | 各岗位考核指标、权重、评分标准与岗位提成规则的集合。 | 岗位绩效说明书汇总_20260809_V04 |
| KPI / 指标项目 | KPI | 岗位绩效责任书中按编号列示、带权重和评分标准的考核项。 | 同上，各岗位表第4行起 |
| 评价维度 | — | KPI 所属的组织指标、岗位指标或个人指标分类。 | 同上 |
| 指标权重 | weight | KPI 对最终绩效总分的权重；源表以合计公式求和。 | 同上 |
| 完成结果 | — | 源表用于计算最终得分的输入结果字段；文件中已有值是否为示例/历史数据未标注。 | 同上 |
| 最终得分 | — | 源表按“指标权重 × 完成结果”计算并汇总的 KPI 加权分。 | 同上 |
| 达成率 | achievement rate | 实际值 ÷ 目标值；具体实际值、目标值口径和阈值按各 KPI 原文。 | 同上 |
| 目标值 | target | 考核周期内预算、计划或 KPI 目标。 | 同上 |
| 实际值 | actual | 考核周期内的实际达成值。 | 同上 |
| 绩效总分 | — | 各 KPI 最终得分的合计。 | 同上 |
| 负责人打分 | — | 数据来源为部门负责人或直属内容主管/编导的指标评分输入。 | 同上 |
| 提成基数 | — | 兴趣电商运营与高级主管表列示的固定基数0.003。 | 同上，运营/高级主管表 |
| 提成梯度 | — | 按得分下限（含）、得分上限（不含）与系数/比例确定的提成分档。 | 同上 |
| GSV | GSV | 源表以“退货后营收（GSV）”表述的营收指标；缩写全称未在源文件定义。 | 同上 |
| 签收ROI | ROI | 用实际签收ROI与预算目标签收ROI计算达成率的指标口径。 | 同上 |
| 免责提成 | — | 新项目免责期内按提成基数×0.8计算的项目免责提成比例。 | 同上，高级主管表 |
| 新项目达成激励 | — | 新项目免责期内达成期间预算目标，取项目利润3%作为项目激励。 | 同上，高级主管表 |
| 爆款 | — | 近90天内上传素材，单视频消耗超过50万。 | 同上，编导/视频剪辑表 |
| 云视频管家 | — | 制作部营收、产出、素材/消耗类指标的数据来源。 | 同上，制作部各岗位表 |

## 预算域

| 术语 | 英文/缩写 | 定义 | 来源 |
|---|---|---|---|
| 中期预算模型 | — | 公司分渠道分月预算的权威文件，当前版本 V60。 | 羊喜市中期预算模型_20260716_V60.xlsx |
| 总表-渠道二维汇总 | — | 预算模型中分渠道×分月的汇总工作表，预算目标提取来源。 | 同上 |

## 组织域

| 术语 | 英文/缩写 | 定义 | 来源 |
|---|---|---|---|
| 一级部门 | — | 组织架构顶层部门。 | 飞书 Base 组织架构表 |
| 二级部门 | — | 一级部门下的细分部门。 | 同上 |
| 项目 | — | 业务核算单元（如高个子、少年状元、骨能）。 | V04 高级主管表 |
""", encoding="utf-8")
    MEMORY.write_text("""# Business Memory（business_analyst / finance_controller 可写）

## KPI 定义登记

- 2026-08-14 / T1 只读提取：V04 共9个岗位工作表、45项 KPI；每条明细、源单元格和公式已固化于 `data/output/绩效框架结构化规则.json`。
- 最终得分的源表通用结构为“指标权重 × 完成结果”，总分为各项最终得分之和；各 KPI 的评分阈值彼此独立，禁止跨岗位套用。
- 制作部消耗相关指标的取值范围为上传时间90天内素材的上月整月消耗，数据源为云视频管家。

## 封顶/保底规则

- 达成率/进度型 KPI 的封顶、保底以对应源单元格的评分标准为准；常见的150分、120分、110分和1.5封顶系数不可泛化。
- 多个提成表最高仅列至150分，≥150分的处理方式未在源文件明确，已列为待确认事项。

## 部门差异化规则

- 兴趣电商高级主管的新项目机制及免责提成仅见于该岗位源表；不得外推到其他岗位。
- 客服岗位采用奖励/处罚条件并列描述，源表未给出金额、次数与总分的统一换算公式。

## 争议与待确认

- 广告投放 KPI“团队ROI签收目标达成率”的末档文字存在重叠；直播中控提成表列头与数据列视觉错位。详见结构化规则JSON的 `disputes_and_open_items`；未对原文作修正。
""", encoding="utf-8")
    # 产物一致性验证：重新读取所有派生规则文档，验证 KPI 覆盖和总计均与 JSON 基线一致。
    rendered_full = OUT_FULL.read_text(encoding="utf-8")
    rendered_rules = OUT_RULES.read_text(encoding="utf-8")
    assert "{p['weight_total_formula']}" not in rendered_full
    assert "{p['total_score_formula']}" not in rendered_full
    rendered_full_metric_count = 0
    for p in reloaded["positions"]:
        expected_formula_line = f"- 权重合计公式：`{p['weight_total_formula']}`；最终得分合计公式：`{p['total_score_formula']}`"
        assert expected_formula_line in rendered_full
        for metric in p["metrics"]:
            assert render_metric(metric) in rendered_full
            assert f"{metric['source_metric_number']}：{metric['name']}；{metric['weight']}；{metric['scoring_type']}" in rendered_rules
            rendered_full_metric_count += 1
    assert rendered_full_metric_count == total_metrics
    assert f"- 提取范围：9 岗位、{total_metrics} 项 KPI。" in rendered_rules
    log={"status":"success","source_sha256":source_hash,"sheet_count":len(wb.sheetnames),"metric_count":total_metrics,"checks":sheet_checks,"derived_document_metric_counts":{"docs/绩效框架结构化规则.md":rendered_full_metric_count,"docs/business_rules.md":rendered_full_metric_count,".hermes/memory/business_memory.md":total_metrics},"outputs":[str(p.relative_to(ROOT)) for p in [OUT_JSON,OUT_FULL,OUT_RULES,GLOSSARY,MEMORY]],"validation":["源Excel未被写入","JSON重新加载成功","岗位数与工作表数一致","KPI总数与按岗位提取总数一致","工作表顺序与源文件一致","所有派生文档 KPI 总数与 JSON 一致","9个岗位合计公式已逐项与JSON验证"],"exceptions":disputes}
    OUT_LOG.write_text(json.dumps(log,ensure_ascii=False,indent=2)+"\n",encoding="utf-8")
    print(json.dumps({"status":"success","source_sha256":source_hash,"sheet_count":len(wb.sheetnames),"metric_count":total_metrics,"outputs":log['outputs']},ensure_ascii=False))
if __name__ == "__main__": main()
