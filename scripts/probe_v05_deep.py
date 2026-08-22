#!/usr/bin/env python3
"""V05 深查：运营 sheet 右侧扩展列、两个新增规则 sheet 全量单元格（只读）。"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "1.reference/绩效模板/岗位绩效说明书汇总_20260821_V05.xlsx"
wb = load_workbook(SOURCE, data_only=False)

ws = wb["兴趣电商-运营"]
print("### 兴趣电商-运营 全部非空单元格（N 列及以后）")
for row in ws.iter_rows(min_col=14):
    for c in row:
        if c.value is not None:
            s = str(c.value).replace("\n", "\\n")
            print(f"{c.coordinate} ({c.data_type}): {s[:120]}")

for name in ["提成比例与绩效工资核算规则", "绩效豁免规则"]:
    ws = wb[name]
    print(f"\n### {name} 全量单元格 rows={ws.max_row} cols={ws.max_column}")
    print("merged:", [str(r) for r in ws.merged_cells.ranges])
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                s = str(c.value).replace("\n", "\\n")
                print(f"{c.coordinate} ({c.data_type}): {s[:200]}")
