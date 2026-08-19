#!/usr/bin/env python3
"""验证绩效结果说明书 V03 的部门、渠道、下钻与可重跑输出。只读 XLSX。"""
from __future__ import annotations

import argparse
import json
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data/output/绩效结果说明书_20260818_V03.xlsx"
BASELINE = ROOT / "data/output/绩效结果说明书_20260818_V03_幂等基线.json"
REPORT = ROOT / "data/output/绩效结果说明书_20260818_V03_验证摘要.json"


def serialise(value: Any) -> Any:
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def workbook_snapshot(path: Path) -> dict[str, list[list[Any]]]:
    wb = load_workbook(path, read_only=True, data_only=False)
    return {
        ws.title: [[serialise(value) for value in row] for row in ws.iter_rows(values_only=True)]
        for ws in wb.worksheets
    }


def validate(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    expected_tail = ["录入明细", "输出明细（文本+计算对照）"]
    if len(wb.sheetnames) < 3 or wb.sheetnames[-2:] != expected_tail:
        raise RuntimeError(f"Sheet 结构不符合 V03 约定：{wb.sheetnames}")
    output = wb[expected_tail[1]]
    headers = [cell.value for cell in next(output.iter_rows(min_row=1, max_row=1))]
    required_output = {"部门", "渠道", "员工ID", "Result记录ID", "月度总分", "加权得分", "提成金额", "V04评分标准原文"}
    missing = required_output - set(headers)
    if missing:
        raise RuntimeError(f"输出明细缺列：{sorted(missing)}")
    rows = [dict(zip(headers, values)) for values in output.iter_rows(min_row=2, values_only=True)]
    if not rows:
        raise RuntimeError("输出明细为空")
    by_department: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        if row["渠道"] in (None, ""):
            raise RuntimeError(f"渠道列存在空值：Result记录ID={row['Result记录ID']}")
        if not row["V04评分标准原文"]:
            raise RuntimeError(f"V04文本缺失：Result记录ID={row['Result记录ID']}")
        by_department[str(row["部门"])].append(row)
    samples = []
    for department in sorted(by_department):
        if department not in wb.sheetnames:
            raise RuntimeError(f"缺少部门 Sheet：{department}")
        selected = sorted(by_department[department], key=lambda r: (str(r["姓名"]), str(r["员工ID"]), str(r["Result记录ID"])))[0]
        # 复算与部门汇总同样采用“期间 + 员工”粒度，防止跨月分数被错误相加。
        employee_rows = [
            r for r in by_department[department]
            if r["期间"] == selected["期间"] and r["员工ID"] == selected["员工ID"]
        ]
        monthly_total_values = sorted({str(r["月度总分"]) for r in employee_rows})
        weighted_sum = round(sum(float(r["加权得分"]) for r in employee_rows if isinstance(r["加权得分"], (int, float))), 4)
        sheet_rows = list(wb[department].iter_rows(min_row=2, values_only=True))
        # 部门汇总列：期间、姓名、部门、岗位、渠道、员工ID、指标数……
        # 汇总粒度为“期间 + 员工”：跨月结果必须分别保留，不能把不同月份静默合并为一行。
        summary_matches = [
            r for r in sheet_rows
            if r[0] == selected["期间"] and r[5] == selected["员工ID"]
        ]
        if len(summary_matches) != 1:
            raise RuntimeError(
                f"部门 Sheet 员工期间汇总行数异常："
                f"{department}/{selected['期间']}/{selected['员工ID']}={len(summary_matches)}"
            )
        samples.append({
            "department": department,
            "employee": selected["姓名"],
            "employee_id": selected["员工ID"],
            "channel": selected["渠道"],
            "department_sheet": department,
            "result_rows": len(employee_rows),
            "monthly_total_values": monthly_total_values,
            "weighted_sum_numeric": weighted_sum,
            "commission_sum": round(sum(float(r["提成金额"]) for r in employee_rows if isinstance(r["提成金额"], (int, float))), 4),
            "recalculation_consistent": len(monthly_total_values) == 1 and weighted_sum == float(employee_rows[0]["月度总分"]),
        })
    return {
        "status": "PASS",
        "sheets": wb.sheetnames,
        "department_count": len(by_department),
        "department_employee_rows": {d: len({r["员工ID"] for r in rs}) for d, rs in by_department.items()},
        "department_result_rows": {d: len(rs) for d, rs in by_department.items()},
        "channel_distribution": dict(Counter(str(r["渠道"]) for r in rows)),
        "input_rows": wb["录入明细"].max_row - 1,
        "output_rows": len(rows),
        "samples": samples,
        "three_department_sampling_possible": len(by_department) >= 3,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--write-baseline", action="store_true")
    parser.add_argument("--compare-baseline", action="store_true")
    args = parser.parse_args()
    result = validate(OUT)
    snapshot = workbook_snapshot(OUT)
    if args.write_baseline:
        BASELINE.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
        result["idempotence"] = "BASELINE_WRITTEN"
    if args.compare_baseline:
        if not BASELINE.exists():
            raise RuntimeError(f"缺少幂等基线：{BASELINE}")
        baseline = json.loads(BASELINE.read_text(encoding="utf-8"))
        if snapshot != baseline:
            raise RuntimeError("重跑前后工作簿单元格值不一致")
        result["idempotence"] = "PASS：重跑前后工作簿单元格值一致"
    REPORT.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
