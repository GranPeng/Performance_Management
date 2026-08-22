#!/usr/bin/env python3
"""核查 V05 岗位 sheet 第 9 行以后内容：提成表/特殊规则是被移除还是表头改名（只读）。"""
from openpyxl import load_workbook
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
wb4 = load_workbook(ROOT.parent / "1.中期预算调整项目/3.outputs/3.岗位绩效/岗位绩效说明书汇总_20260809_V04_最新绩效框架.xlsx", data_only=False)
wb5 = load_workbook(ROOT / "1.reference/绩效模板/岗位绩效说明书汇总_20260821_V05.xlsx", data_only=False)

for name in ["兴趣电商-运营", "兴趣电商-高级主管", "兴趣电商-主播", "兴趣电商-客服", "制作部-编导", "制作部-视频剪辑", "制作部-摄影师", "兴趣电商-广告投放", "兴趣电商-直播中控"]:
    for tag, wb in [("V04", wb4), ("V05", wb5)]:
        ws = wb[name]
        nonempty = []
        for row in ws.iter_rows(min_row=9):
            for c in row:
                if c.value is not None:
                    s = str(c.value)[:50].replace("\n", "\\n")
                    nonempty.append(f"{c.coordinate}={s}")
        print(f"--- {name} [{tag}] rows>8 非空 {len(nonempty)} 个")
        if tag == "V05" or name in ("兴趣电商-高级主管", "兴趣电商-运营"):
            for x in nonempty:
                print("   ", x)
