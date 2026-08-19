#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""提取 V04 直播中控 + 广告投放 sheet 完整内容（核对 D-013/D-015 落文依据）"""
import openpyxl

SRC = '/Users/gran/Library/Mobile Documents/com~apple~CloudDocs/工作空间/工作文件WIP/1.进行中/1.中期预算调整项目/3.outputs/3.岗位绩效/岗位绩效说明书汇总_20260809_V04_最新绩效框架.xlsx'

wb = openpyxl.load_workbook(SRC, data_only=False)

for sn in ['兴趣电商-直播中控', '兴趣电商-广告投放']:
    if sn not in wb.sheetnames:
        print(f'!! sheet 不存在: {sn}')
        continue
    ws = wb[sn]
    print(f'\n{"="*80}\n### SHEET: {sn}  dims={ws.dimensions}\n{"="*80}')
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        cells = []
        for c in row:
            if c.value is not None and str(c.value).strip() != '':
                cells.append(f'{c.coordinate}={c.value!r}')
        if cells:
            print(' | '.join(cells))
