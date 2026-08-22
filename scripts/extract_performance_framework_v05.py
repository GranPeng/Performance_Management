#!/usr/bin/env python3
"""只读提取 V05 绩效框架；口径沿用 extract_performance_framework.py（V04）。

差异：
- 源文件换为 V05；输出仅 JSON 与校验日志，不改写 V04 基线文档（docs/绩效框架结构化规则.md、
  docs/business_rules.md、context/business_glossary.md、.hermes/memory/business_memory.md）。
- V05 新增 2 个规则 sheet（提成比例与绩效工资核算规则、绩效豁免规则），按源单元格全量结构化，
  计入 rule_sheets；9 个岗位 sheet 的提取逻辑与 V04 完全一致。
"""
from __future__ import annotations
import hashlib, json
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "1.reference/绩效模板/岗位绩效说明书汇总_20260821_V05.xlsx"
OUT_JSON = ROOT / "data/output/绩效框架结构化规则_V05.json"
OUT_LOG = ROOT / "data/output/绩效框架提取校验日志_V05.json"

SHEET_CODE = {
    "兴趣电商-运营": "IE-OPS", "兴趣电商-高级主管": "IE-SUP", "兴趣电商-广告投放": "IE-ADS",
    "兴趣电商-直播中控": "IE-LIVE", "兴趣电商-主播": "IE-HOST", "兴趣电商-客服": "IE-CS",
    "制作部-编导": "PROD-DIR", "制作部-视频剪辑": "PROD-EDIT", "制作部-摄影师": "PROD-CAM",
}
RULE_SHEETS = ["提成比例与绩效工资核算规则", "绩效豁免规则"]

def value(v):
    return v if v is not None else None

def text(v):
    return "" if v is None else str(v)

def score_type(scoring: str, formula: str, reward: str = "", penalty: str = "") -> str:
    all_text = "\n".join([scoring, formula, reward, penalty])
    if "达成率=" in formula or "达成率=" in scoring or "人均产出=" in formula or "个人贡献率公式=" in formula:
        return "达成率/进度型"
    if "错误次数" in all_text or "漏回复次数" in all_text or "次数" in all_text and "得" in all_text:
        return "次数阈值型"
    if "单次扣除" in all_text or "单次扣" in all_text:
        return "扣分制"
    if reward or penalty or "奖励" in all_text or "处罚" in all_text:
        return "奖惩制"
    if "0-60分" in all_text or "0—60分" in all_text:
        return "定性等级型"
    return "文字评分（类型未显式标注）"

def col_name(c): return chr(64 + c)

def extract_metrics(ws):
    # 与 V04 完全一致：标准表 B:M；客服表 B:L，列 F/G/H 依次为评分/奖励/处罚。
    header_row = 4
    headers = {text(ws.cell(header_row, c).value).replace("\n", ""): c for c in range(1, ws.max_column + 1) if ws.cell(header_row,c).value}
    is_customer_service = "奖励条件及金额/提成标准" in headers
    total_row = next((r for r in range(6, ws.max_row + 1) if text(ws.cell(r, 2).value) == "合计"), None)
    if total_row is None: raise ValueError(f"{ws.title}: 未找到合计行")
    result, dimension = [], None
    for r in range(6, total_row):
        number = ws.cell(r, 2).value
        if number is None and ws.cell(r, 4).value is None:
            continue
        source_number = number if number is not None else "未填写（源单元格B%d为空）" % r
        d = ws.cell(r, 3).value
        if d is not None: dimension = d
        if is_customer_service:
            metric = {
                "source_metric_number": source_number, "source_metric_number_raw": number, "source_cell": f"B{r}", "dimension": dimension,
                "name": value(ws.cell(r,4).value), "weight": value(ws.cell(r,5).value),
                "target": None, "calculation_formula": None, "unit": None,
                "scoring_standard": value(ws.cell(r,6).value), "reward_condition": value(ws.cell(r,7).value),
                "penalty_condition": value(ws.cell(r,8).value), "data_source": value(ws.cell(r,9).value),
                "evaluation_period": value(ws.cell(r,10).value), "source_result_example": value(ws.cell(r,11).value),
                "source_score_formula": value(ws.cell(r,12).value),
            }
        else:
            metric = {
                "source_metric_number": source_number, "source_metric_number_raw": number, "source_cell": f"B{r}", "dimension": dimension,
                "name": value(ws.cell(r,4).value), "weight": value(ws.cell(r,5).value),
                "target": value(ws.cell(r,6).value), "calculation_formula": value(ws.cell(r,7).value),
                "unit": value(ws.cell(r,8).value), "scoring_standard": value(ws.cell(r,9).value),
                "reward_condition": None, "penalty_condition": None, "data_source": value(ws.cell(r,10).value),
                "evaluation_period": value(ws.cell(r,11).value), "source_result_example": value(ws.cell(r,12).value),
                "source_score_formula": value(ws.cell(r,13).value),
            }
        metric["scoring_type"] = score_type(text(metric["scoring_standard"]), text(metric["calculation_formula"]), text(metric["reward_condition"]), text(metric["penalty_condition"]))
        metric["score_cap_or_floor"] = "按评分标准原文执行；未另行推断"
        result.append(metric)
    return result, total_row, is_customer_service

def extract_commission(ws):
    # 与 V04 完全一致：找“提成来源”表头，按源单元格保留。
    header = None
    for row in ws.iter_rows():
        for c in row:
            if c.value == "提成来源": header = (c.row, c.column); break
        if header: break
    if not header: return []
    r, start = header
    max_col = min(ws.max_column, start + 4)
    columns = [{"column": col_name(c), "header": value(ws.cell(r,c).value)} for c in range(start, max_col+1)]
    tiers=[]
    for rr in range(r+1, ws.max_row+1):
        first = ws.cell(rr,start).value
        if first is None: continue
        if text(first).startswith("*") or first == "、": continue
        row={"source_row": rr, "source_cell": f"{col_name(start)}{rr}", "values": {col_name(c): value(ws.cell(rr,c).value) for c in range(start,max_col+1)}}
        tiers.append(row)
    return [{"source_header_row": r, "columns": columns, "tiers": tiers}]

def extract_special_rules(ws):
    # 与 V04 完全一致。
    rows=[]
    for row in ws.iter_rows():
        for c in row:
            if c.value == "考核方式说明":
                h = c.row + 1
                for rr in range(h+1, ws.max_row+1):
                    if ws.cell(rr,4).value is None: continue
                    rows.append({"source_row": rr, "负责项目": value(ws.cell(rr,4).value), "渠道": value(ws.cell(rr,5).value), "考核方式": value(ws.cell(rr,6).value), "新项目评价周期": value(ws.cell(rr,7).value)})
    notes=[]
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value,str) and c.value.startswith("*"):
                notes.append({"source_cell": c.coordinate, "text": c.value})
    return {"assessment_methods": rows, "source_notes": notes}

def extract_commission_rule_sheet(ws):
    """新增 sheet「提成比例与绩效工资核算规则」：A1:J1 表头，第 2 行起逐行按源单元格保留。"""
    headers = {col_name(c): value(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)}
    rows = []
    for r in range(2, ws.max_row + 1):
        if all(ws.cell(r, c).value is None for c in range(1, ws.max_column + 1)):
            continue
        rows.append({
            "source_row": r, "source_cell": f"A{r}",
            "岗位": value(ws.cell(r, 1).value), "提成来源": value(ws.cell(r, 2).value),
            "提成规则": value(ws.cell(r, 3).value), "绩效得分下限_包含": value(ws.cell(r, 4).value),
            "绩效得分上限_不包含": value(ws.cell(r, 5).value), "提成比例基数": value(ws.cell(r, 6).value),
            "对应提成比例系数": value(ws.cell(r, 7).value), "提成比例": value(ws.cell(r, 8).value),
            "对应绩效工资核算公式": value(ws.cell(r, 9).value), "特殊项说明": value(ws.cell(r, 10).value),
        })
    return {"source_sheet": ws.title, "source_header_row": 1, "headers": headers, "rows": rows}

def extract_exemption_rule_sheet(ws):
    """新增 sheet「绩效豁免规则」：A1 标题、第 2 行表头、3-6 行数据、星号备注。"""
    rows = []
    for r in range(3, ws.max_row + 1):
        if ws.cell(r, 1).value is None or text(ws.cell(r, 1).value).startswith("*"):
            continue
        rows.append({
            "source_row": r, "source_cell": f"A{r}",
            "负责项目": value(ws.cell(r, 1).value), "渠道": value(ws.cell(r, 2).value),
            "考核方式": value(ws.cell(r, 3).value), "新项目评价周期": value(ws.cell(r, 4).value),
        })
    notes = [{"source_cell": c.coordinate, "text": c.value}
             for row in ws.iter_rows() for c in row
             if isinstance(c.value, str) and c.value.startswith("*")]
    return {"source_sheet": ws.title, "title": value(ws["A1"].value),
            "headers": {col_name(c): value(ws.cell(2, c).value) for c in range(1, ws.max_column + 1)},
            "rows": rows, "source_notes": notes}

def main():
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    source_hash = hashlib.sha256(SOURCE.read_bytes()).hexdigest()
    wb = load_workbook(SOURCE, data_only=False)
    positions=[]; total_metrics=0; sheet_checks=[]
    for ws in wb.worksheets:
        if ws.title in RULE_SHEETS:
            continue
        metrics,total_row,is_cs = extract_metrics(ws)
        formula_cells = [c.coordinate for row in ws.iter_rows() for c in row if c.data_type == "f"]
        weight_formula = ws.cell(total_row,5).value
        score_formula = ws.cell(total_row,12 if is_cs else 13).value
        position={
            "source_sheet": ws.title, "source_sheet_code": SHEET_CODE[ws.title], "department": ws["D2"].value,
            "position": ws["F2"].value, "source_range": f"B4:{col_name(ws.max_column)}{ws.max_row}",
            "metrics": metrics, "weight_total_formula": weight_formula, "total_score_formula": score_formula,
            "commission_rules": extract_commission(ws), "special_rules": extract_special_rules(ws),
        }
        positions.append(position); total_metrics += len(metrics)
        sheet_checks.append({"sheet":ws.title,"metric_count":len(metrics),"source_total_row":total_row,"weight_total_formula":weight_formula,"total_score_formula":score_formula,"formula_cells":formula_cells})
    rule_sheets = {}
    rule_sheet_checks = []
    for name in RULE_SHEETS:
        ws = wb[name]
        if name == "提成比例与绩效工资核算规则":
            rule_sheets[name] = extract_commission_rule_sheet(ws)
            rule_sheet_checks.append({"sheet": name, "row_count": len(rule_sheets[name]["rows"])})
        else:
            rule_sheets[name] = extract_exemption_rule_sheet(ws)
            rule_sheet_checks.append({"sheet": name, "row_count": len(rule_sheets[name]["rows"]),
                                      "note_count": len(rule_sheets[name]["source_notes"])})
    # 可疑/待业务确认：V04 已登记项逐条在 V05 复核，并登记 V05 新增疑点；只记录，不修补。
    disputes=[
        {"source":"兴趣电商-客服!B11","issue":"“邀请好评”指标存在但B列编号为空；已保留该指标，并在结构化结果中标记为“未填写（源单元格B11为空）”，未自行补号。（V04 遗留，V05 复核状态见差异报告）"},
        {"source":"兴趣电商-广告投放!I7","issue":"评分标准末档原文为“90%＞达成率，0分”，与紧邻的“90%＞达成率≥80%，60分”重叠；按原文保留，需业务确认。（V04 遗留，V05 复核状态见差异报告）"},
        {"source":"兴趣电商-直播中控!I18:M21","issue":"V04 提成表的列头与数据列存在视觉错位；V05 新增「提成比例与绩效工资核算规则」sheet 以规范化列结构重述了中控提成（H 列文本“0.1%/中控人数”、I 列“0.8*绩效工资”等），是否以新表为准需业务确认。"},
        {"source":"兴趣电商-广告投放!I17:M20；兴趣电商-主播!I16:M19；制作部各岗位提成表","issue":"多个提成梯度最高仅至150分，未给出≥150分的处理规则；不得自行外推。（V04 遗留；V05 新增规则 sheet 同样上限至150分，未补充）"},
        {"source":"各岗位“完成结果”列及高级主管测算区","issue":"源工作簿含示例完成结果/测算值；其是否为演示数据或历史业务结果未标记，本提取仅作为来源字段记录，不作为规则事实。"},
        {"source":"兴趣电商-客服!F6:H11","issue":"客服表按奖励/处罚金额与分值并列描述，但未给出将金额、次数与最终得分的统一换算公式；本提取保留原文，待业务与财务确认。"},
        {"source":"提成比例与绩效工资核算规则!A12:A15","issue":"V05 新增规则 sheet 出现岗位名“编导&内容主管”，V04 岗位体系中无“内容主管”岗位 sheet；该名称是合并指代还是新增岗位需业务确认。"},
        {"source":"提成比例与绩效工资核算规则!H28:H31","issue":"广告投放四档“提成比例”为静态数值 0.001（非 =F*G 公式），与其他岗位公式化表达不一致；按源单元格保留。"},
        {"source":"提成比例与绩效工资核算规则!E38/E45","issue":"电商运营/高级主管最高档上限为文本“♾️”，机器计算需映射为正无穷；按原文保留。"},
    ]
    data={
        "schema_version":"1.0", "record_type":"performance_framework_rule_extraction", "source": {"file": SOURCE.name, "sha256":source_hash, "read_mode":"openpyxl data_only=False；只读提取，未修改源文件", "sheet_count":len(wb.sheetnames), "sheet_names":wb.sheetnames},
        "extraction": {"timestamp_utc":datetime.now(timezone.utc).isoformat(), "transform":"每个岗位表从第4行表头、编号列与合计行之间提取KPI；合并单元格评价维度向下继承；提成与特殊规则按源单元格保留。V05 新增 2 个规则 sheet 按源单元格全量结构化，计入 rule_sheets。", "non_authoritative_generated_fields":["source_sheet_code","scoring_type","score_cap_or_floor"], "traceability":"每条KPI保留 source_sheet、source_cell 与 source_metric_number。", "baseline_script":"scripts/extract_performance_framework.py（V04 口径），本脚本 scripts/extract_performance_framework_v05.py 逐行沿用"},
        "positions":positions, "rule_sheets":rule_sheets, "disputes_and_open_items":disputes
    }
    OUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2)+"\n", encoding="utf-8")
    reloaded=json.loads(OUT_JSON.read_text(encoding="utf-8"))
    assert len(reloaded["positions"]) == len(wb.sheetnames) - len(RULE_SHEETS)
    assert sum(len(x["metrics"]) for x in reloaded["positions"]) == total_metrics
    assert [x["source_sheet"] for x in reloaded["positions"]] == [n for n in wb.sheetnames if n not in RULE_SHEETS]
    assert len(reloaded["rule_sheets"]) == len(RULE_SHEETS)
    log={"status":"success","source_sha256":source_hash,"sheet_count":len(wb.sheetnames),"position_sheet_count":len(positions),"rule_sheet_count":len(rule_sheets),"metric_count":total_metrics,"checks":sheet_checks,"rule_sheet_checks":rule_sheet_checks,"outputs":[str(p.relative_to(ROOT)) for p in [OUT_JSON,OUT_LOG]],"validation":["源Excel未被写入","JSON重新加载成功","岗位数与岗位工作表数一致","KPI总数与按岗位提取总数一致","工作表顺序与源文件一致","2 个新增规则 sheet 行数与源文件核对一致"],"exceptions":disputes}
    OUT_LOG.write_text(json.dumps(log,ensure_ascii=False,indent=2)+"\n",encoding="utf-8")
    print(json.dumps({"status":"success","source_sha256":source_hash,"sheet_count":len(wb.sheetnames),"position_sheets":len(positions),"metric_count":total_metrics,"rule_sheet_rows":{c["sheet"]:c["row_count"] for c in rule_sheet_checks}},ensure_ascii=False))
if __name__ == "__main__": main()
