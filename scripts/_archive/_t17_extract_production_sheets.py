#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""提取 V04 源文件制作部三个 sheet 的完整单元格内容（含公式文本）"""
import openpyxl, json, hashlib, sys

SRC = '/Users/gran/Library/Mobile Documents/com~apple~CloudDocs/工作空间/工作文件WIP/1.进行中/1.中期预算调整项目/3.outputs/3.岗位绩效/岗位绩效说明书汇总_20260809_V04_最新绩效框架.xlsx'

with open(SRC, 'rb') as f:
    sha = hashlib.sha256(f.read()).hexdigest()
print('SHA-256:', sha)

wb = openpyxl.load_workbook(SRC, data_only=False)

sheets = ['制作部-编导', '制作部-视频剪辑', '制作部-摄影师']
for sn in sheets:
    if sn not in wb.sheetnames:
        print(f'!! sheet 不存在: {sn}')
        continue
    ws = wb[sn]
    print(f'\n{"="*80}\n### SHEET: {sn}  dims={ws.dimensions}  max_row={ws.max_row} max_col={ws.max_column}\n{"="*80}')
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        cells = []
        for c in row:
            if c.value is not None and str(c.value).strip() != '':
                cells.append(f'{c.coordinate}={c.value!r}')
        if cells:
            print(' | '.join(cells))
