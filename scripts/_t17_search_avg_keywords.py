#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""搜索 V04 全簿中与小组平均值/均额/平均相关的单元格，定位「小组平均值」公式原文"""
import openpyxl

SRC = '/Users/gran/Library/Mobile Documents/com~apple~CloudDocs/工作空间/工作文件WIP/1.进行中/1.中期预算调整项目/3.outputs/3.岗位绩效/岗位绩效说明书汇总_20260809_V04_最新绩效框架.xlsx'

wb = openpyxl.load_workbook(SRC, data_only=False)

KEYWORDS = ['平均', '均额', '小组', '组', '人均', '均分', '平均数']
for sn in wb.sheetnames:
    ws = wb[sn]
    hits = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for c in row:
            v = c.value
            if v is None:
                continue
            s = str(v)
            for kw in KEYWORDS:
                if kw in s:
                    hits.append(f'  {c.coordinate} [{kw}]: {s[:120]!r}')
                    break
    if hits:
        print(f'\n### SHEET: {sn}  ({len(hits)} hits)')
        for h in hits:
            print(h)
