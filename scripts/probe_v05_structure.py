#!/usr/bin/env python3
"""V05 源文件结构探查（只读）。"""
from openpyxl import load_workbook
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "1.reference/绩效模板/岗位绩效说明书汇总_20260821_V05.xlsx"

wb = load_workbook(SOURCE, data_only=False)
print("sheet_count:", len(wb.sheetnames))
for ws in wb.worksheets:
    print("=" * 60)
    print("SHEET:", repr(ws.title), "rows=", ws.max_row, "cols=", ws.max_column)
    print("D2=", repr(ws["D2"].value), "F2=", repr(ws["F2"].value))
    # 打印前 8 行非空单元格概览
    for r in range(1, min(ws.max_row, 8) + 1):
        cells = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None:
                s = str(v).replace("\n", "\\n")
                cells.append(f"{ws.cell(r, c).coordinate}={s[:40]}")
        if cells:
            print(f"  row{r}: " + " | ".join(cells))
